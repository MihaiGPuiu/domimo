from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponseNotAllowed
from django.views.decorators.http import require_POST
from django.contrib.auth import authenticate, login
from django.contrib import messages
from .form import UserRegistrationForm
import pandas as pd
import json
import logging
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import Apartment

from .models import Apartment, Account  # Import your Django models
from .form import AnnouncementForm

# Configure logger
logger = logging.getLogger(__name__)

# Read the Excel file
excel_file = 'apartment.xlsx'
df = pd.read_excel(excel_file)

# Add a unique identifier to each row if not already present
if 'id' not in df.columns:
    df['id'] = range(1, len(df) + 1)

def display_columns(request):
    """
    View to display filtered and paginated list of apartments.
    """
    query = request.GET.get("search", "")
    filter_title = request.GET.get("filter_title", "")
    filter_city = request.GET.get("filter_city", "")
    filter_surface = request.GET.get("filter_surface", "")

    # Initial filtered DataFrame (start with the entire dataset)
    filtered_df = df.copy()

    # Perform filtering based on the query and filters
    if query:
        filtered_df = filtered_df[filtered_df['Title'].str.contains(query, case=False, na=False)]
    if filter_title:
        filtered_df = filtered_df[filtered_df['Title'].str.contains(filter_title, case=False, na=False)]
    if filter_city:
        filtered_df = filtered_df[filtered_df['City'].str.contains(filter_city, case=False, na=False)]
    if filter_surface:
        filtered_df = filtered_df[filtered_df['Surface'].str.contains(filter_surface, case=False, na=False)]

    # Check if there are no results and prepare the message
    no_results_message = None
    if filtered_df.empty:
        no_results_message = f"Nu s-au gasit rezultate pentru '{query}'"
        filtered_df = df.copy()  # Reset to show all apartments

    # Convert DataFrame to list of dictionaries
    apartments_list = []
    for index, row in filtered_df.iterrows():
        apartment = {
            'id': row['id'],
            'title': row['Title'],
            'city': row['City'],
            'surface': row['Surface'],
            'price': row['Price'],
            'image': row['Image_links'],
        }
        apartments_list.append(apartment)

    # Paginator initialization
    p = Paginator(apartments_list, 15)  # Initialize Paginator with filtered list
    page_number = request.GET.get('page')
    pages = p.get_page(page_number)

    cities = df['City'].unique()

    return render(request, '1.html', {
        'apartments': pages,
        'query': query,
        'filter_title': filter_title,
        'filter_city': filter_city,
        'cities': cities,
        'no_results_message': no_results_message,
    })

def favorites(request):
    """
    View for displaying favorite apartments (currently empty).
    """
    return render(request, 'favorites.html', {'apartments': []})
@require_POST
def toggle_favorite(request, apartment_id):
    if request.method == 'POST':
        # Fetch the apartment object based on apartment_id
        apartment = get_object_or_404(Apartment, id=apartment_id)
        
        # Toggle the favorite status
        apartment.is_favorite = not apartment.is_favorite
        apartment.save()
        
        # Return a JsonResponse indicating success
        return JsonResponse({'success': True, 'is_favorite': apartment.is_favorite})
    
    # Return an error response if the request method is not POST
    return JsonResponse({'error': 'Invalid request method'}, status=405)
def favorite_list(request):
    favorites = request.user.favorite_apartments.all()
    return render(request, 'favorites.html', {'favorites': favorites})
def home(request):
    """
    View for rendering the home page.
    """
    return render(request, 'home.html')

def read_excel_file():
    """
    Function to read the Excel file initially.
    """
    try:
        df = pd.read_excel(excel_file)
    except FileNotFoundError:
        df = pd.DataFrame(columns=['Title', 'City', 'Surface', 'Price', 'Image_links'])  # Create empty DataFrame if file does not exist
    return df

def save_to_excel(title, type_, city, district, surface, price, image_links, location):
    """
    Function to save form data to the Excel file.
    """
    file_path = 'apartment.xlsx'
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        logger.error(f"File not found: {file_path}")
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title='Announcements')
        worksheet.append(['Price', 'Title', 'Type', 'City', 'District', 'Surface', 'Image_links', 'Location'])
    except Exception as e:
        logger.error(f"Error loading workbook: {e}")
        return

    worksheet = workbook.active
    try:
        worksheet.append([price, title, type_, city, district, surface, image_links, location])
        workbook.save(file_path)
        logger.info(f"Data saved to Excel: {price}, {title}, {type_}, {city}, {district}, {surface}, {image_links}, {location}")
    except Exception as e:
        logger.error(f"Error appending data to worksheet: {e}")

def form_view(request):
    """
    View for rendering the form page.
    """
    context = {}
    return render(request, 'form.html', context)

def submit_form(request):
    """
    View for handling form submission and saving data to the Excel file.
    """
    if request.method == 'POST':
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            title = form.cleaned_data['title']
            type_ = form.cleaned_data['type']
            city = form.cleaned_data['city']
            district = form.cleaned_data['district']
            surface = form.cleaned_data['surface']
            price = form.cleaned_data['price']
            image_links = form.cleaned_data['image_links']
            location = form.cleaned_data['location']
            logger.info(f"Form data: {price}, {title}, {type_}, {city}, {district}, {surface}, {image_links}, {location}")
            save_to_excel(title, type_, city, district, surface, price, image_links, location)
            return redirect('home')
        else:
            logger.warning(f"Form is invalid: {form.errors}")
    else:
        form = AnnouncementForm()

    apartments = read_excel_file()
    return render(request, 'form.html', {'form': form, 'apartments': apartments})

@require_POST
def ajax_is_favorite(request):
    """
    AJAX view to mark an apartment as favorite.
    """
    if not request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return HttpResponseNotAllowed(['POST'])
    
    data = json.loads(request.body)
    object_id = data.get('id')
    favorite_object = get_object_or_404(Apartment, id=object_id)
    favorite_object.is_favorite = True
    favorite_object.save()
    
    return JsonResponse({"success": True})

def login_view(request):
    """
    View for handling user login.
    """
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, "Invalid username or password.")
            return redirect('login')
    else:
        return render(request, "login.html")
def register(request):
    if request.method== "POST":
        form = UserRegistrationForm(require_POST)
        if form.is_valid():
            user=form.save()
            login(request,user)
            messages.success(request,"Registration successful")
            return redirect('home')
    